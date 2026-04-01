"""
Add comprehensive graphs for All Predictions sheet to Excel file
"""

import pandas as pd
import matplotlib.pyplot as plt
import numpy as np
from sklearn.metrics import r2_score
from scipy import stats
from openpyxl import load_workbook
from openpyxl.drawing.image import Image
import os

# Specify the Excel file
excel_file = 'chemical_adsorbent_report_20260223_122015.xlsx'

print("=" * 60)
print("Creating Graphs for All Predictions Sheet")
print("=" * 60)

# Read the All Predictions sheet
print(f"\n[1] Reading '{excel_file}'...")
predictions_df = pd.read_excel(excel_file, sheet_name='All Predictions')
print(f"Loaded {len(predictions_df)} predictions")

# Create comprehensive visualizations
print("\n[2] Creating comprehensive visualizations...")
fig = plt.figure(figsize=(20, 14))

# 1. Actual vs Predicted Scatter Plot
plt.subplot(3, 4, 1)
scatter = plt.scatter(predictions_df['actual_removal_percentage (%)'], 
                      predictions_df['predicted_removal_percentage (%)'], 
                      alpha=0.6, s=50, c=predictions_df['absolute_error (%)'], cmap='RdYlGn_r')
plt.plot([predictions_df['actual_removal_percentage (%)'].min(), 
          predictions_df['actual_removal_percentage (%)'].max()], 
         [predictions_df['actual_removal_percentage (%)'].min(), 
          predictions_df['actual_removal_percentage (%)'].max()], 
         'r--', lw=2, label='Perfect Prediction')
plt.xlabel('Actual Removal %', fontsize=11)
plt.ylabel('Predicted Removal %', fontsize=11)
plt.title('Actual vs Predicted (All Data)', fontsize=12, fontweight='bold')
plt.colorbar(scatter, label='Absolute Error (%)')
plt.legend()
plt.grid(True, alpha=0.3)

# 2. Error Distribution
plt.subplot(3, 4, 2)
plt.hist(predictions_df['error (%)'], bins=40, color='steelblue', alpha=0.7, edgecolor='black')
plt.axvline(x=0, color='red', linestyle='--', linewidth=2, label='Zero Error')
plt.xlabel('Error (%)', fontsize=11)
plt.ylabel('Frequency', fontsize=11)
plt.title('Prediction Error Distribution', fontsize=12, fontweight='bold')
plt.legend()
plt.grid(True, alpha=0.3)

# 3. Absolute Error Distribution
plt.subplot(3, 4, 3)
plt.hist(predictions_df['absolute_error (%)'], bins=40, color='coral', alpha=0.7, edgecolor='black')
plt.xlabel('Absolute Error (%)', fontsize=11)
plt.ylabel('Frequency', fontsize=11)
plt.title('Absolute Error Distribution', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 4. Box Plot - Actual vs Predicted
plt.subplot(3, 4, 4)
plt.boxplot([predictions_df['actual_removal_percentage (%)'], 
             predictions_df['predicted_removal_percentage (%)']], 
            tick_labels=['Actual', 'Predicted'], patch_artist=True,
            boxprops=dict(facecolor='lightblue', color='blue'),
            medianprops=dict(color='red', linewidth=2))
plt.ylabel('Removal Percentage (%)', fontsize=11)
plt.title('Actual vs Predicted - Box Plot', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3, axis='y')

# 5. Residual Plot
plt.subplot(3, 4, 5)
plt.scatter(predictions_df['predicted_removal_percentage (%)'], 
            predictions_df['error (%)'], 
            alpha=0.6, color='purple', s=50)
plt.axhline(y=0, color='red', linestyle='--', linewidth=2)
plt.xlabel('Predicted Removal %', fontsize=11)
plt.ylabel('Residual Error (%)', fontsize=11)
plt.title('Residual Plot', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 6. Cumulative Error Plot
plt.subplot(3, 4, 6)
sorted_abs_error = np.sort(predictions_df['absolute_error (%)'])
cumulative_pct = np.arange(1, len(sorted_abs_error) + 1) / len(sorted_abs_error) * 100
plt.plot(sorted_abs_error, cumulative_pct, linewidth=2, color='darkgreen')
plt.xlabel('Absolute Error (%)', fontsize=11)
plt.ylabel('Cumulative Percentage', fontsize=11)
plt.title('Cumulative Error Distribution', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 7. Top 20 Predictions Comparison
plt.subplot(3, 4, 7)
n_samples = min(20, len(predictions_df))
indices = range(n_samples)
width = 0.35
plt.bar([i - width/2 for i in indices], 
        predictions_df['actual_removal_percentage (%)'].iloc[:n_samples], 
        width, label='Actual', alpha=0.8, color='blue')
plt.bar([i + width/2 for i in indices], 
        predictions_df['predicted_removal_percentage (%)'].iloc[:n_samples], 
        width, label='Predicted', alpha=0.8, color='orange')
plt.xlabel('Sample Index', fontsize=11)
plt.ylabel('Removal Percentage (%)', fontsize=11)
plt.title(f'First {n_samples} Predictions Comparison', fontsize=12, fontweight='bold')
plt.legend()
plt.xticks(indices, [str(i) for i in indices], rotation=45)
plt.grid(True, alpha=0.3, axis='y')

# 8. Error by Actual Value
plt.subplot(3, 4, 8)
plt.scatter(predictions_df['actual_removal_percentage (%)'], 
            predictions_df['absolute_error (%)'], 
            alpha=0.6, color='crimson', s=50)
plt.xlabel('Actual Removal %', fontsize=11)
plt.ylabel('Absolute Error (%)', fontsize=11)
plt.title('Error by Actual Value', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 9. Q-Q Plot for Errors
plt.subplot(3, 4, 9)
stats.probplot(predictions_df['error (%)'], dist="norm", plot=plt)
plt.title('Q-Q Plot (Error Normality)', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 10. Line Plot - First 50 samples
plt.subplot(3, 4, 10)
n_line = min(50, len(predictions_df))
x_line = range(n_line)
plt.plot(x_line, predictions_df['actual_removal_percentage (%)'].iloc[:n_line], 
         'o-', label='Actual', linewidth=2, markersize=6, color='blue')
plt.plot(x_line, predictions_df['predicted_removal_percentage (%)'].iloc[:n_line], 
         's-', label='Predicted', linewidth=2, markersize=5, color='orange')
plt.xlabel('Sample Index', fontsize=11)
plt.ylabel('Removal Percentage (%)', fontsize=11)
plt.title(f'Trend Comparison (First {n_line} Samples)', fontsize=12, fontweight='bold')
plt.legend()
plt.grid(True, alpha=0.3)

# 11. Percentage Error Distribution
plt.subplot(3, 4, 11)
percentage_error = (predictions_df['error (%)'] / predictions_df['actual_removal_percentage (%)']) * 100
percentage_error = percentage_error.replace([np.inf, -np.inf], np.nan).dropna()
plt.hist(percentage_error, bins=40, color='teal', alpha=0.7, edgecolor='black')
plt.xlabel('Percentage Error (%)', fontsize=11)
plt.ylabel('Frequency', fontsize=11)
plt.title('Relative Error Distribution', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 12. Statistical Summary Text
plt.subplot(3, 4, 12)
plt.axis('off')
r2 = r2_score(predictions_df['actual_removal_percentage (%)'], 
              predictions_df['predicted_removal_percentage (%)'])
rmse = np.sqrt(np.mean(predictions_df['error (%)']**2))

stats_text = f"""
Prediction Statistics:

Total Predictions: {len(predictions_df)}

Actual Range: 
  {predictions_df['actual_removal_percentage (%)'].min():.2f} - {predictions_df['actual_removal_percentage (%)'].max():.2f}%

Predicted Range: 
  {predictions_df['predicted_removal_percentage (%)'].min():.2f} - {predictions_df['predicted_removal_percentage (%)'].max():.2f}%

Mean Absolute Error: 
  {predictions_df['absolute_error (%)'].mean():.3f}%

Median Abs Error: 
  {predictions_df['absolute_error (%)'].median():.3f}%

Max Abs Error: 
  {predictions_df['absolute_error (%)'].max():.3f}%

RMSE: {rmse:.3f}%

R² Score: {r2:.4f}
"""
plt.text(0.1, 0.5, stats_text, fontsize=10, family='monospace', 
         verticalalignment='center', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
plt.title('Statistical Summary', fontsize=12, fontweight='bold')

plt.tight_layout()
graph_file = 'all_predictions_graphs_overview.png'
plt.savefig(graph_file, dpi=300, bbox_inches='tight')
print(f"Saved: {graph_file}")
plt.close()

# Create detailed scatter plot
print("\n[3] Creating detailed scatter plot...")
fig1, ax1 = plt.subplots(figsize=(12, 10))
scatter = ax1.scatter(predictions_df['actual_removal_percentage (%)'], 
                      predictions_df['predicted_removal_percentage (%)'], 
                      alpha=0.6, s=100, c=predictions_df['absolute_error (%)'], 
                      cmap='RdYlGn_r', edgecolors='black', linewidth=0.5)
ax1.plot([predictions_df['actual_removal_percentage (%)'].min(), 
          predictions_df['actual_removal_percentage (%)'].max()], 
         [predictions_df['actual_removal_percentage (%)'].min(), 
          predictions_df['actual_removal_percentage (%)'].max()], 
         'r--', lw=3, label='Perfect Prediction Line')
ax1.set_xlabel('Actual Removal Percentage (%)', fontsize=14, fontweight='bold')
ax1.set_ylabel('Predicted Removal Percentage (%)', fontsize=14, fontweight='bold')
ax1.set_title('Detailed: Actual vs Predicted Removal Percentage\n(All Predictions)', 
              fontsize=16, fontweight='bold', pad=20)
cbar = plt.colorbar(scatter, ax=ax1)
cbar.set_label('Absolute Error (%)', fontsize=12, fontweight='bold')
ax1.legend(fontsize=12)
ax1.grid(True, alpha=0.3, linestyle='--')
plt.tight_layout()
scatter_file = 'all_predictions_scatter.png'
plt.savefig(scatter_file, dpi=300, bbox_inches='tight')
print(f"Saved: {scatter_file}")
plt.close()

# Create detailed bar chart
print("\n[4] Creating detailed bar chart...")
fig2, ax2 = plt.subplots(figsize=(16, 8))
n_bars = min(len(predictions_df), len(predictions_df))  # Show all predictions
indices = range(n_bars)
width = 0.35
bars1 = ax2.bar([i - width/2 for i in indices], 
                predictions_df['actual_removal_percentage (%)'].iloc[:n_bars], 
                width, label='Actual', alpha=0.8, color='royalblue', edgecolor='black')
bars2 = ax2.bar([i + width/2 for i in indices], 
                predictions_df['predicted_removal_percentage (%)'].iloc[:n_bars], 
                width, label='Predicted', alpha=0.8, color='darkorange', edgecolor='black')
ax2.set_xlabel('Sample Index', fontsize=14, fontweight='bold')
ax2.set_ylabel('Removal Percentage (%)', fontsize=14, fontweight='bold')
ax2.set_title(f'Detailed: Actual vs Predicted Values (All {n_bars} Samples)', 
              fontsize=16, fontweight='bold', pad=20)
ax2.legend(fontsize=12, loc='best')
ax2.set_xticks(indices)
ax2.set_xticklabels([str(i) for i in indices], rotation=90, fontsize=8)
ax2.grid(True, alpha=0.3, axis='y', linestyle='--')
plt.tight_layout()
bar_file = 'all_predictions_bars.png'
plt.savefig(bar_file, dpi=300, bbox_inches='tight')
print(f"Saved: {bar_file}")
plt.close()

# Add graphs to Excel file
print("\n[5] Adding graphs to Excel file...")
wb = load_workbook(excel_file)

# Remove existing "All Predictions Graphs" sheet if it exists
if 'All Predictions Graphs' in wb.sheetnames:
    del wb['All Predictions Graphs']

# Create new sheet for graphs
ws_graphs = wb.create_sheet('All Predictions Graphs', 2)  # Insert after Visualizations

# Add the overview graph
img1 = Image(graph_file)
img1.width = 1600
img1.height = 1120
ws_graphs.add_image(img1, 'A1')

# Add the scatter plot
img2 = Image(scatter_file)
img2.width = 960
img2.height = 800
ws_graphs.add_image(img2, 'A62')

# Add the bar chart
img3 = Image(bar_file)
img3.width = 1280
img3.height = 640
ws_graphs.add_image(img3, 'Q62')

# Save the Excel file
wb.save(excel_file)
wb.close()

print(f"\n✓ Successfully added 'All Predictions Graphs' sheet to {excel_file}")

# Clean up temporary image files
os.remove(graph_file)
os.remove(scatter_file)
os.remove(bar_file)
print("✓ Cleaned up temporary files")

print("\n" + "=" * 60)
print("Graph Generation Complete!")
print("=" * 60)
print(f"\nOpen '{excel_file}' and navigate to the")
print("'All Predictions Graphs' sheet to view the visualizations.")
print("\nThe sheet contains:")
print("  1. Overview dashboard (12 different plots)")
print("  2. Detailed scatter plot")
print("  3. Detailed bar comparison (all samples)")
