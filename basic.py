"""
Chemical Adsorbent Prediction using Linear Regression
Features: all columns from the input Excel file except the target
Target: removal percentage
"""

# Import necessary libraries
import numpy as np
import pandas as pd
from sklearn.model_selection import train_test_split
from sklearn.linear_model import LinearRegression
from sklearn.metrics import mean_squared_error, r2_score, mean_absolute_error
from sklearn.preprocessing import OneHotEncoder
from sklearn.compose import ColumnTransformer
from sklearn.pipeline import Pipeline
from sklearn.impute import SimpleImputer
import matplotlib.pyplot as plt
import seaborn as sns

# Set random seed for reproducibility
np.random.seed(42)

# Business rule: removal/error percentages must not exceed 99.0
MAX_PERCENT = 99.0
PERCENT_DECIMALS = 1


def clamp_percent(values):
    """Clip percentages to [0, MAX_PERCENT] and round to configured decimals."""
    return np.round(np.clip(values, 0, MAX_PERCENT), PERCENT_DECIMALS)

print("=" * 60)
print("Chemical Adsorbent Prediction System")
print("=" * 60)

print("\n[1] Loading Excel data...")
excel_file = "adsorption dataset.xlsx"

try:
    df = pd.read_excel(excel_file)
except FileNotFoundError as exc:
    raise FileNotFoundError(
        f"Input file not found: {excel_file}. Place it in the project folder."
    ) from exc

# Standardize column names and target column
df.columns = df.columns.str.strip()
df = df.loc[:, ~df.columns.str.contains(r"^Unnamed", na=False)]
target_col_original = "% Removal"
if target_col_original not in df.columns:
    raise ValueError(
        f"Target column '{target_col_original}' not found. Columns: {list(df.columns)}"
    )

# Map adsorbent codes to full names if present
adsorbent_map = {
    "A1": "Activated Carbon",
    "A2": "Biochar",
    "A3": "Chitosan composite",
}
if "Adsorbent" in df.columns:
    df["Adsorbent"] = df["Adsorbent"].replace(adsorbent_map)

df.rename(columns={target_col_original: "removal_percentage (%)"}, inplace=True)
target_col = "removal_percentage (%)"

print(f"Dataset shape: {df.shape}")
print("\nFirst 5 rows of the dataset:")
print(df.head())

# Statistical summary
print("\n[2] Statistical Summary:")
print(df.describe(include="all"))

# Check for missing values
print("\n[3] Checking for missing values:")
print(df.isnull().sum())

# Separate features and target
# Drop target leakage columns that are derived from removal percentage
leakage_cols = {"Ce (mg/L)", "qe (mg/g)"}
feature_cols = [col for col in df.columns if col != target_col and col not in leakage_cols]
X = df[feature_cols]
y = df[target_col]
input_feature_names = feature_cols

print(f"\nFeatures (X) shape: {X.shape}")
print(f"Target (y) shape: {y.shape}")

# Split data into training and testing sets (80-20 split)
print("\n[4] Splitting data into training and testing sets (80-20)...")
X_train, X_test, y_train, y_test = train_test_split(
    X, y, test_size=0.2, random_state=42
)

print(f"Training set size: {X_train.shape[0]} samples")
print(f"Testing set size: {X_test.shape[0]} samples")

# Create and train Linear Regression model
print("\n[5] Training Linear Regression model...")
categorical_cols = X.select_dtypes(include=["object", "string"]).columns.tolist()
numeric_cols = [col for col in X.columns if col not in categorical_cols]

categorical_transformer = Pipeline(
    steps=[
        ("imputer", SimpleImputer(strategy="most_frequent")),
        ("onehot", OneHotEncoder(handle_unknown="ignore")),
    ]
)

numeric_transformer = Pipeline(
    steps=[
        ("imputer", SimpleImputer(strategy="median")),
    ]
)

preprocessor = ColumnTransformer(
    transformers=[
        ("cat", categorical_transformer, categorical_cols),
        ("num", numeric_transformer, numeric_cols),
    ]
)

model = Pipeline(
    steps=[
        ("preprocessor", preprocessor),
        ("regressor", LinearRegression()),
    ]
)
model.fit(X_train, y_train)
print("Model training completed!")

# Display model coefficients
print("\n[6] Model Coefficients:")
feature_names = []
if categorical_cols:
    cat_names = model.named_steps["preprocessor"].named_transformers_["cat"].named_steps["onehot"].get_feature_names_out(categorical_cols)
    feature_names.extend(cat_names)
feature_names.extend(numeric_cols)
coefficients = pd.DataFrame({
    'Feature': feature_names,
    'Coefficient': model.named_steps["regressor"].coef_
})
print(coefficients)
print(f"\nIntercept: {model.named_steps['regressor'].intercept_:.4f}")

# Make predictions
print("\n[7] Making predictions...")
y_train_pred = model.predict(X_train)
y_test_pred = model.predict(X_test)

# Clip predictions to valid range [0, 99.0] for removal percentage
print(f"\n[7a] Clipping predictions to valid range (0-{MAX_PERCENT:.1f}%)...")
y_train_pred = clamp_percent(y_train_pred)
y_test_pred = clamp_percent(y_test_pred)
print(f"  Training predictions range: {y_train_pred.min():.2f}% - {y_train_pred.max():.2f}%")
print(f"  Testing predictions range: {y_test_pred.min():.2f}% - {y_test_pred.max():.2f}%")

# Evaluate model performance
print("\n[8] Model Performance:")
print("\nTraining Set Metrics:")
train_r2 = r2_score(y_train, y_train_pred)
train_mse = mean_squared_error(y_train, y_train_pred)
train_rmse = np.sqrt(train_mse)
train_mae = mean_absolute_error(y_train, y_train_pred)

print(f"  R² Score: {train_r2:.4f}")
print(f"  Mean Squared Error (MSE): {train_mse:.4f}")
print(f"  Root Mean Squared Error (RMSE): {train_rmse:.4f}")
print(f"  Mean Absolute Error (MAE): {train_mae:.4f}")

print("\nTesting Set Metrics:")
test_r2 = r2_score(y_test, y_test_pred)
test_mse = mean_squared_error(y_test, y_test_pred)
test_rmse = np.sqrt(test_mse)
test_mae = mean_absolute_error(y_test, y_test_pred)

print(f"  R² Score: {test_r2:.4f}")
print(f"  Mean Squared Error (MSE): {test_mse:.4f}")
print(f"  Root Mean Squared Error (RMSE): {test_rmse:.4f}")
print(f"  Mean Absolute Error (MAE): {test_mae:.4f}")

# Sample predictions
print("\n[9] Sample Predictions:")
sample_predictions = pd.DataFrame({
    'Actual': y_test[:10].values,
    'Predicted': y_test_pred[:10],
    'Difference': y_test[:10].values - y_test_pred[:10]
})
print(sample_predictions)

# Full predictions for all rows (same structure as input with extra columns)
print("\n[9b] Building full prediction output for all metal combinations...")
all_predictions = X.copy()
all_predictions['actual_removal_percentage (%)'] = y.values
# Clip predictions to valid range [0, 99.0]
raw_predictions = model.predict(X)
all_predictions['predicted_removal_percentage (%)'] = clamp_percent(raw_predictions)
all_predictions['error (%)'] = (
    all_predictions['actual_removal_percentage (%)'] -
    all_predictions['predicted_removal_percentage (%)']
)
# Round and cap errors at 99.0%
all_predictions['error (%)'] = np.round(all_predictions['error (%)'], PERCENT_DECIMALS)
all_predictions['absolute_error (%)'] = np.round(
    np.minimum(np.abs(all_predictions['error (%)']), MAX_PERCENT),
    PERCENT_DECIMALS,
)

print(f"  Total predictions: {len(all_predictions)}")
if 'Metal' in all_predictions.columns:
    print(f"  Unique metals in predictions: {all_predictions['Metal'].unique()}")
if 'Adsorbent' in all_predictions.columns:
    print(f"  Unique adsorbents: {all_predictions['Adsorbent'].unique()}")
print(f"  Predicted range: {all_predictions['predicted_removal_percentage (%)'].min():.2f}% - {all_predictions['predicted_removal_percentage (%)'].max():.2f}%")
print(f"  Max absolute error: {all_predictions['absolute_error (%)'].max():.2f}%")

# Generate predictions for all adsorbent-metal combinations
print("\n[9c] Generating predictions for all adsorbent-metal combinations...")

# Define target metals for prediction (specified metals)
target_metals = ['Pb', 'Cd', 'Hg', 'As', 'Cr', 'Cu', 'Ni', 'Zn']
print(f"  Target metals: {', '.join(target_metals)}")

# Get unique adsorbents from the dataset
adsorbents = sorted(df['Adsorbent'].unique().tolist())
print(f"  Adsorbents: {', '.join(adsorbents)}")

# Create template from the dataset (use representative parameter values)
# Get common parameter combinations from the dataset
parameter_sets = df[['Dosage (g/L)', 'Temp (°C)', 'pH', 'Time (min)', 'RPM', 'C0 (mg/L)']].drop_duplicates()

# Generate all combinations
metal_combinations = []
for adsorbent in adsorbents:
    for metal in target_metals:
        for _, params in parameter_sets.iterrows():
            combination = {
                'Adsorbent': adsorbent,
                'Metal': metal,
                'Dosage (g/L)': params['Dosage (g/L)'],
                'Temp (°C)': params['Temp (°C)'],
                'pH': params['pH'],
                'Time (min)': params['Time (min)'],
                'RPM': params['RPM'],
                'C0 (mg/L)': params['C0 (mg/L)']
            }
            metal_combinations.append(combination)

# Create DataFrame with all combinations
metal_predictions_df = pd.DataFrame(metal_combinations)

# Make predictions for all combinations
print(f"  Generating {len(metal_predictions_df)} predictions...")
raw_metal_predictions = model.predict(metal_predictions_df)
# Clip predictions to valid range [0, 99.0]
metal_predictions_df['predicted_removal_percentage (%)'] = clamp_percent(raw_metal_predictions)

# For these predictions, we don't have actual values, so we note them as synthetic
metal_predictions_df['prediction_type'] = 'Synthetic (All Metal Combinations)'

print(f"  ✓ Generated predictions for {len(target_metals)} metals × {len(adsorbents)} adsorbents")
print(f"  ✓ Total combinations: {len(metal_predictions_df)}")
print(f"  ✓ Predicted range: {metal_predictions_df['predicted_removal_percentage (%)'].min():.2f}% - {metal_predictions_df['predicted_removal_percentage (%)'].max():.2f}%")

# Show summary by metal
print("\n  Metal-wise prediction summary:")
for metal in target_metals:
    metal_data = metal_predictions_df[metal_predictions_df['Metal'] == metal]
    print(f"    {metal:>3} - Mean: {metal_data['predicted_removal_percentage (%)'].mean():.1f}%, "
          f"Range: {metal_data['predicted_removal_percentage (%)'].min():.1f}%-{metal_data['predicted_removal_percentage (%)'].max():.1f}%")

# Visualizations
print("\n[10] Creating visualizations...")

# Create figure with multiple subplots
fig = plt.figure(figsize=(16, 12))

# 1. Actual vs Predicted (Training)
plt.subplot(3, 3, 1)
plt.scatter(y_train, y_train_pred, alpha=0.5, color='blue', label='Training')
plt.plot([y_train.min(), y_train.max()], [y_train.min(), y_train.max()], 
         'r--', lw=2, label='Perfect Prediction')
plt.xlabel('Actual Removal Percentage (%)')
plt.ylabel('Predicted Removal Percentage (%)')
plt.title('Training Set: Actual vs Predicted')
plt.legend()
plt.grid(True, alpha=0.3)

# 2. Actual vs Predicted (Testing)
plt.subplot(3, 3, 2)
plt.scatter(y_test, y_test_pred, alpha=0.5, color='green', label='Testing')
plt.plot([y_test.min(), y_test.max()], [y_test.min(), y_test.max()], 
         'r--', lw=2, label='Perfect Prediction')
plt.xlabel('Actual Removal Percentage (%)')
plt.ylabel('Predicted Removal Percentage (%)')
plt.title('Testing Set: Actual vs Predicted')
plt.legend()
plt.grid(True, alpha=0.3)

# 3. Residuals (Testing)
plt.subplot(3, 3, 3)
residuals = y_test - y_test_pred
plt.scatter(y_test_pred, residuals, alpha=0.5, color='purple')
plt.axhline(y=0, color='r', linestyle='--', lw=2)
plt.xlabel('Predicted Removal Percentage (%)')
plt.ylabel('Residuals (%)')
plt.title('Residual Plot (Testing Set)')
plt.grid(True, alpha=0.3)

# 4. Feature Importance (Coefficients)
plt.subplot(3, 3, 4)
plt.barh(feature_names, model.named_steps["regressor"].coef_, color='skyblue')
plt.xlabel('Coefficient Value')
plt.title('Feature Importance (Coefficients)')
plt.grid(True, alpha=0.3)

# 5. Distribution of Actual vs Predicted
plt.subplot(3, 3, 5)
plt.hist(y_test, bins=30, alpha=0.5, label='Actual', color='blue')
plt.hist(y_test_pred, bins=30, alpha=0.5, label='Predicted', color='orange')
plt.xlabel('Removal Percentage (%)')
plt.ylabel('Frequency')
plt.title('Distribution: Actual vs Predicted')
plt.legend()
plt.grid(True, alpha=0.3)

# 6. Residuals Distribution
plt.subplot(3, 3, 6)
plt.hist(residuals, bins=30, color='purple', alpha=0.7, edgecolor='black')
plt.xlabel('Residuals (%)')
plt.ylabel('Frequency')
plt.title('Residuals Distribution')
plt.grid(True, alpha=0.3)

# 7. Feature Correlation Heatmap
plt.subplot(3, 3, 7)
numeric_df = df.select_dtypes(include=[np.number])
correlation_matrix = numeric_df.corr()
sns.heatmap(correlation_matrix, annot=True, fmt='.2f', cmap='coolwarm', 
            square=True, cbar_kws={'shrink': 0.8})
plt.title('Feature Correlation Heatmap')

# 8. Q-Q Plot (Residuals Normality)
plt.subplot(3, 3, 8)
from scipy import stats
stats.probplot(residuals, dist="norm", plot=plt)
plt.title('Q-Q Plot (Residuals Normality)')
plt.grid(True, alpha=0.3)

# 9. Prediction Error
plt.subplot(3, 3, 9)
prediction_error = np.abs(y_test - y_test_pred)
plt.scatter(range(len(prediction_error)), prediction_error, alpha=0.5, color='red')
plt.xlabel('Sample Index')
plt.ylabel('Absolute Error (%)')
plt.title('Prediction Error Distribution')
plt.grid(True, alpha=0.3)

plt.tight_layout()
plt.savefig('chemical_adsorbent_analysis.png', dpi=300, bbox_inches='tight')
print("Visualizations saved as 'chemical_adsorbent_analysis.png'")
plt.close()  # Close the plot instead of showing it

# Build graph analysis summary for the Excel report
residual_mean = float(np.mean(residuals))
residual_std = float(np.std(residuals))
error_mean = float(np.mean(prediction_error))
error_median = float(np.median(prediction_error))

top_corr_text = "N/A"
if target_col in correlation_matrix.columns:
    corr_with_target = correlation_matrix[target_col].drop(labels=[target_col])
    if not corr_with_target.empty:
        top_corr = corr_with_target.abs().sort_values(ascending=False).head(3)
        top_corr_text = "; ".join([f"{idx}: {corr_with_target[idx]:.3f}" for idx in top_corr.index])

analysis_rows = [
    {"Plot": "Actual vs Predicted (Train)", "Insight": f"R2: {train_r2:.3f} | RMSE: {train_rmse:.3f}"},
    {"Plot": "Actual vs Predicted (Test)", "Insight": f"R2: {test_r2:.3f} | RMSE: {test_rmse:.3f}"},
    {"Plot": "Residuals (Test)", "Insight": f"Mean: {residual_mean:.3f} | Std: {residual_std:.3f}"},
    {"Plot": "Residuals Distribution", "Insight": f"Median abs error: {error_median:.3f}"},
    {"Plot": "Prediction Error", "Insight": f"Mean abs error: {error_mean:.3f}"},
    {"Plot": "Correlation Heatmap", "Insight": f"Top correlations with target: {top_corr_text}"},
]

# Example: Make a single prediction with custom input
print("\n[11] Example: Making prediction with custom input...")
custom_input = X.head(1).copy()

print("\nInput features:")
print(custom_input)

custom_prediction = model.predict(custom_input)
custom_prediction = clamp_percent(custom_prediction)
print(f"\nPredicted Removal Percentage: {custom_prediction[0]:.1f}%")

# Save model (optional)
print("\n[12] Saving model...")
import pickle

with open('linear_regression_model.pkl', 'wb') as f:
    pickle.dump(model, f)
print("Model saved as 'linear_regression_model.pkl'")

# Save dataset to CSV
df.to_csv('chemical_adsorbent_data.csv', index=False)
print("Dataset saved as 'chemical_adsorbent_data.csv'")

# Save full predictions to CSV
all_predictions.to_csv('chemical_adsorbent_predictions.csv', index=False)
print("Predictions saved as 'chemical_adsorbent_predictions.csv'")

# Save metal combinations predictions to CSV
metal_predictions_df.to_csv('metal_combinations_predictions.csv', index=False)
print("Metal combinations predictions saved as 'metal_combinations_predictions.csv'")
print(f"  ✓ Contains predictions for {len(target_metals)} metals: {', '.join(target_metals)}")

def generate_excel_report(excel_path):
    from openpyxl import load_workbook
    from openpyxl.drawing.image import Image
    from openpyxl.styles import Font, PatternFill, Alignment

    with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
        # Sheet 1: Complete Dataset
        df.to_excel(writer, sheet_name='Dataset', index=False)

        # Sheet 2: Training Data
        train_data = X_train.copy()
        train_data['removal_percentage (%)'] = y_train.values
        train_data.to_excel(writer, sheet_name='Training Data', index=False)

        # Sheet 3: Testing Data with Predictions
        test_results = X_test.copy()
        test_results['actual_removal_percentage (%)'] = y_test.values
        test_results['predicted_removal_percentage (%)'] = y_test_pred
        test_results['error (%)'] = np.round(y_test.values - y_test_pred, PERCENT_DECIMALS)
        # Cap absolute error at 99.0% (business rule)
        test_results['absolute_error (%)'] = np.round(
            np.minimum(np.abs(test_results['error (%)']), MAX_PERCENT),
            PERCENT_DECIMALS,
        )
        test_results.to_excel(writer, sheet_name='Testing & Predictions', index=False)

        # Sheet 3b: All Predictions (same as input with predictions)
        all_predictions.to_excel(writer, sheet_name='All Predictions', index=False)

        # Sheet 3c: Metal Combinations Predictions (auto from dataset)
        metal_predictions_df.to_excel(writer, sheet_name='Metal Combinations', index=False)

        # Sheet 4: Model Performance Metrics
        metrics_data = {
            'Metric': ['R² Score', 'Mean Squared Error (MSE)', 'Root Mean Squared Error (RMSE)', 'Mean Absolute Error (MAE)'],
            'Training Set': [train_r2, train_mse, train_rmse, train_mae],
            'Testing Set': [test_r2, test_mse, test_rmse, test_mae]
        }
        metrics_df = pd.DataFrame(metrics_data)
        metrics_df.to_excel(writer, sheet_name='Model Metrics', index=False)

        # Sheet 5: Model Coefficients
        coef_data = pd.DataFrame({
            'Feature': feature_names,
            'Coefficient': model.named_steps["regressor"].coef_,
            'Abs_Coefficient': np.abs(model.named_steps["regressor"].coef_)
        })
        coef_data = coef_data.sort_values('Abs_Coefficient', ascending=False)
        coef_data.to_excel(writer, sheet_name='Coefficients', index=False)

        # Sheet 6: Statistical Summary
        summary_df = df.describe()
        summary_df.to_excel(writer, sheet_name='Statistical Summary')

        # Sheet 7: Correlation Matrix
        corr_df = numeric_df.corr()
        corr_df.to_excel(writer, sheet_name='Correlation Matrix')

        # Sheet 8: Graph Analysis
        analysis_df = pd.DataFrame(analysis_rows)
        analysis_df.to_excel(writer, sheet_name='Graph Analysis', index=False)

    wb = load_workbook(excel_path)

    # Create a new sheet for the graph
    if 'Visualizations' in wb.sheetnames:
        del wb['Visualizations']
    ws_graph = wb.create_sheet('Visualizations', 0)

    # Add the graph image
    img = Image('chemical_adsorbent_analysis.png')
    img.width = 1600
    img.height = 1200
    ws_graph.add_image(img, 'A1')

    # Format the Model Metrics sheet
    ws_metrics = wb['Model Metrics']
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_font = Font(color='FFFFFF', bold=True, size=12)

    for cell in ws_metrics[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Format the Coefficients sheet
    ws_coef = wb['Coefficients']
    for cell in ws_coef[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Format the Dataset sheet
    ws_data = wb['Dataset']
    for cell in ws_data[1]:
        cell.fill = PatternFill(start_color='70AD47', end_color='70AD47', fill_type='solid')
        cell.font = Font(color='FFFFFF', bold=True, size=11)
        cell.alignment = Alignment(horizontal='center', vertical='center')

    # Auto-adjust column widths for all sheets
    for sheet_name in wb.sheetnames:
        if sheet_name != 'Visualizations':
            ws = wb[sheet_name]
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                for cell in column:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 20)
                ws.column_dimensions[column_letter].width = adjusted_width

    # Add a summary sheet at the beginning
    if 'Summary' in wb.sheetnames:
        del wb['Summary']
    ws_summary = wb.create_sheet('Summary', 1)

    # Add summary information
    ws_summary['A1'] = 'Chemical Adsorbent Prediction Report'
    ws_summary['A1'].font = Font(size=18, bold=True, color='1F4E78')
    ws_summary['A3'] = 'Project Information'
    ws_summary['A3'].font = Font(size=14, bold=True)
    ws_summary['A4'] = 'Algorithm:'
    ws_summary['B4'] = 'Linear Regression'
    ws_summary['A5'] = 'Date:'
    ws_summary['B5'] = 'March 7, 2026'
    ws_summary['A6'] = 'Total Samples:'
    ws_summary['B6'] = len(df)
    ws_summary['A7'] = 'Training Samples:'
    ws_summary['B7'] = len(X_train)
    ws_summary['A8'] = 'Testing Samples:'
    ws_summary['B8'] = len(X_test)

    ws_summary['A10'] = 'Model Performance (Testing Set)'
    ws_summary['A10'].font = Font(size=14, bold=True)
    ws_summary['A11'] = 'R² Score:'
    ws_summary['B11'] = round(test_r2, 4)
    ws_summary['A12'] = 'RMSE:'
    ws_summary['B12'] = round(test_rmse, 4)
    ws_summary['A13'] = 'MAE:'
    ws_summary['B13'] = round(test_mae, 4)

    ws_summary['A15'] = 'Features Used'
    ws_summary['A15'].font = Font(size=14, bold=True)
    for i, feature in enumerate(input_feature_names, start=16):
        ws_summary[f'A{i}'] = f"{i-15}. {feature.title()}"

    ws_summary['A22'] = 'Target Variable'
    ws_summary['A22'].font = Font(size=14, bold=True)
    ws_summary['A23'] = 'Removal Percentage (%)'

    ws_summary['A25'] = 'Metal Combinations Predicted'
    ws_summary['A25'].font = Font(size=14, bold=True)
    ws_summary['A26'] = f"Metals: {', '.join(target_metals)}"
    ws_summary['A27'] = f'Total Combinations: {len(metal_predictions_df)}'
    ws_summary['A28'] = 'Note: Predicted removal/error capped at 0-99.0%'

    # Set column widths for summary
    ws_summary.column_dimensions['A'].width = 35
    ws_summary.column_dimensions['B'].width = 20

    wb.save(excel_path)


# Create Excel file with multiple sheets and embedded graphs
print("\n[13] Creating comprehensive Excel report with embedded graphs...")
excel_file = 'chemical_adsorbent_report.xlsx'
try:
    generate_excel_report(excel_file)
except PermissionError:
    timestamp = pd.Timestamp.now().strftime("%Y%m%d_%H%M%S")
    excel_file = f"chemical_adsorbent_report_{timestamp}.xlsx"
    generate_excel_report(excel_file)
print(f"Excel report saved as '{excel_file}'")
print("  - Sheet 1: Summary")
print("  - Sheet 2: Visualizations (with embedded graphs)")
print("  - Sheet 3: Dataset")
print("  - Sheet 4: Training Data")
print("  - Sheet 5: Testing & Predictions")
print("  - Sheet 6: All Predictions")
print("  - Sheet 7: Metal Combinations (all adsorbent-metal combinations)")
print("  - Sheet 8: Model Metrics")
print("  - Sheet 9: Coefficients")
print("  - Sheet 10: Statistical Summary")
print("  - Sheet 11: Correlation Matrix")
print("  - Sheet 12: Graph Analysis")

print("\n" + "=" * 60)
print("Analysis Complete!")
print("=" * 60)
print(f"\n📊 Excel Report: {excel_file}")
print(f"📈 Visualizations: chemical_adsorbent_analysis.png")
print(f"💾 Model File: linear_regression_model.pkl")
print(f"📄 CSV Data: chemical_adsorbent_data.csv")
print("📄 CSV Predictions: chemical_adsorbent_predictions.csv")
print("📄 CSV Metal Combinations: metal_combinations_predictions.csv")

# Create detailed graphs for All Predictions sheet
print("\n[14] Creating detailed graphs for All Predictions...")

# Read the All Predictions sheet from the Excel file
predictions_df = pd.read_excel(excel_file, sheet_name='All Predictions')

# Create comprehensive visualizations
fig = plt.figure(figsize=(20, 14))

# 1. Actual vs Predicted Scatter Plot
plt.subplot(3, 4, 1)
plt.scatter(predictions_df['actual_removal_percentage (%)'], 
            predictions_df['predicted_removal_percentage (%)'], 
            alpha=0.6, s=50, c=predictions_df['absolute_error (%)'], cmap='RdYlGn_r')
plt.plot([predictions_df['actual_removal_percentage (%)'].min(), 
          predictions_df['actual_removal_percentage (%)'].max()], 
         [predictions_df['actual_removal_percentage (%)'].min(), 
          predictions_df['actual_removal_percentage (%)'].max()], 
         'r--', lw=2, label='Perfect Prediction')
plt.xlabel('Actual Removal %', fontsize=11)
plt.ylabel('Predicted Removal %', fontsize=11)
plt.title('Actual vs Predicted (All Data)', fontsize=12, fontweight='bold')
plt.colorbar(label='Absolute Error (%)')
plt.legend()
plt.grid(True, alpha=0.3)

# 2. Error Distribution
plt.subplot(3, 4, 2)
plt.hist(predictions_df['error (%)'], bins=40, color='steelblue', alpha=0.7, edgecolor='black')
plt.axvline(x=0, color='red', linestyle='--', linewidth=2, label='Zero Error')
plt.xlabel('Error (%)', fontsize=11)
plt.ylabel('Frequency', fontsize=11)
plt.title('Prediction Error Distribution', fontsize=12, fontweight='bold')
plt.legend()
plt.grid(True, alpha=0.3)

# 3. Absolute Error Distribution
plt.subplot(3, 4, 3)
plt.hist(predictions_df['absolute_error (%)'], bins=40, color='coral', alpha=0.7, edgecolor='black')
plt.xlabel('Absolute Error (%)', fontsize=11)
plt.ylabel('Frequency', fontsize=11)
plt.title('Absolute Error Distribution', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 4. Box Plot - Actual vs Predicted
plt.subplot(3, 4, 4)
plt.boxplot([predictions_df['actual_removal_percentage (%)'], 
             predictions_df['predicted_removal_percentage (%)']], 
            labels=['Actual', 'Predicted'], patch_artist=True,
            boxprops=dict(facecolor='lightblue', color='blue'),
            medianprops=dict(color='red', linewidth=2))
plt.ylabel('Removal Percentage (%)', fontsize=11)
plt.title('Actual vs Predicted - Box Plot', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3, axis='y')

# 5. Residual Plot
plt.subplot(3, 4, 5)
plt.scatter(predictions_df['predicted_removal_percentage (%)'], 
            predictions_df['error (%)'], 
            alpha=0.6, color='purple', s=50)
plt.axhline(y=0, color='red', linestyle='--', linewidth=2)
plt.xlabel('Predicted Removal %', fontsize=11)
plt.ylabel('Residual Error (%)', fontsize=11)
plt.title('Residual Plot', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 6. Cumulative Error Plot
plt.subplot(3, 4, 6)
sorted_abs_error = np.sort(predictions_df['absolute_error (%)'])
cumulative_pct = np.arange(1, len(sorted_abs_error) + 1) / len(sorted_abs_error) * 100
plt.plot(sorted_abs_error, cumulative_pct, linewidth=2, color='darkgreen')
plt.xlabel('Absolute Error (%)', fontsize=11)
plt.ylabel('Cumulative Percentage', fontsize=11)
plt.title('Cumulative Error Distribution', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 7. Top 20 Predictions Comparison
plt.subplot(3, 4, 7)
n_samples = min(20, len(predictions_df))
indices = range(n_samples)
width = 0.35
plt.bar([i - width/2 for i in indices], 
        predictions_df['actual_removal_percentage (%)'].iloc[:n_samples], 
        width, label='Actual', alpha=0.8, color='blue')
plt.bar([i + width/2 for i in indices], 
        predictions_df['predicted_removal_percentage (%)'].iloc[:n_samples], 
        width, label='Predicted', alpha=0.8, color='orange')
plt.xlabel('Sample Index', fontsize=11)
plt.ylabel('Removal Percentage (%)', fontsize=11)
plt.title(f'First {n_samples} Predictions Comparison', fontsize=12, fontweight='bold')
plt.legend()
plt.xticks(indices, [str(i) for i in indices], rotation=45)
plt.grid(True, alpha=0.3, axis='y')

# 8. Error by Actual Value
plt.subplot(3, 4, 8)
plt.scatter(predictions_df['actual_removal_percentage (%)'], 
            predictions_df['absolute_error (%)'], 
            alpha=0.6, color='crimson', s=50)
plt.xlabel('Actual Removal %', fontsize=11)
plt.ylabel('Absolute Error (%)', fontsize=11)
plt.title('Error by Actual Value', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 9. Q-Q Plot for Errors
plt.subplot(3, 4, 9)
from scipy import stats
stats.probplot(predictions_df['error (%)'], dist="norm", plot=plt)
plt.title('Q-Q Plot (Error Normality)', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 10. Line Plot - First 50 samples
plt.subplot(3, 4, 10)
n_line = min(50, len(predictions_df))
x_line = range(n_line)
plt.plot(x_line, predictions_df['actual_removal_percentage (%)'].iloc[:n_line], 
         'o-', label='Actual', linewidth=2, markersize=6, color='blue')
plt.plot(x_line, predictions_df['predicted_removal_percentage (%)'].iloc[:n_line], 
         's-', label='Predicted', linewidth=2, markersize=5, color='orange')
plt.xlabel('Sample Index', fontsize=11)
plt.ylabel('Removal Percentage (%)', fontsize=11)
plt.title(f'Trend Comparison (First {n_line} Samples)', fontsize=12, fontweight='bold')
plt.legend()
plt.grid(True, alpha=0.3)

# 11. Percentage Error Distribution
plt.subplot(3, 4, 11)
percentage_error = (predictions_df['error (%)'] / predictions_df['actual_removal_percentage (%)']) * 100
percentage_error = percentage_error.replace([np.inf, -np.inf], np.nan).dropna()
percentage_error = np.clip(percentage_error, -MAX_PERCENT, MAX_PERCENT)
plt.hist(percentage_error, bins=40, color='teal', alpha=0.7, edgecolor='black')
plt.xlabel('Percentage Error (%)', fontsize=11)
plt.ylabel('Frequency', fontsize=11)
plt.title('Relative Error Distribution', fontsize=12, fontweight='bold')
plt.grid(True, alpha=0.3)

# 12. Statistical Summary Text
plt.subplot(3, 4, 12)
plt.axis('off')
stats_text = f"""
Prediction Statistics:

Total Predictions: {len(predictions_df)}

Actual Range: {predictions_df['actual_removal_percentage (%)'].min():.2f} - {predictions_df['actual_removal_percentage (%)'].max():.2f}%

Predicted Range: {predictions_df['predicted_removal_percentage (%)'].min():.2f} - {predictions_df['predicted_removal_percentage (%)'].max():.2f}%

Mean Absolute Error: {predictions_df['absolute_error (%)'].mean():.3f}%

Median Abs Error: {predictions_df['absolute_error (%)'].median():.3f}%

Max Abs Error: {predictions_df['absolute_error (%)'].max():.3f}%

RMSE: {np.sqrt(np.mean(predictions_df['error (%)']**2)):.3f}%

R² Score: {r2_score(predictions_df['actual_removal_percentage (%)'], predictions_df['predicted_removal_percentage (%)']):.4f}
"""
plt.text(0.1, 0.5, stats_text, fontsize=11, family='monospace', 
         verticalalignment='center', bbox=dict(boxstyle='round', facecolor='wheat', alpha=0.5))
plt.title('Statistical Summary', fontsize=12, fontweight='bold')

plt.tight_layout()
plt.savefig('all_predictions_graphs.png', dpi=300, bbox_inches='tight')
print("All Predictions graphs saved as 'all_predictions_graphs.png'")
plt.close()

# Create individual detailed graphs for better visibility
# Graph 1: Large Actual vs Predicted
fig1, ax1 = plt.subplots(figsize=(12, 10))
scatter = ax1.scatter(predictions_df['actual_removal_percentage (%)'], 
                      predictions_df['predicted_removal_percentage (%)'], 
                      alpha=0.6, s=100, c=predictions_df['absolute_error (%)'], 
                      cmap='RdYlGn_r', edgecolors='black', linewidth=0.5)
ax1.plot([predictions_df['actual_removal_percentage (%)'].min(), 
          predictions_df['actual_removal_percentage (%)'].max()], 
         [predictions_df['actual_removal_percentage (%)'].min(), 
          predictions_df['actual_removal_percentage (%)'].max()], 
         'r--', lw=3, label='Perfect Prediction Line')
ax1.set_xlabel('Actual Removal Percentage (%)', fontsize=14, fontweight='bold')
ax1.set_ylabel('Predicted Removal Percentage (%)', fontsize=14, fontweight='bold')
ax1.set_title('Detailed: Actual vs Predicted Removal Percentage\n(All Predictions)', 
              fontsize=16, fontweight='bold', pad=20)
cbar = plt.colorbar(scatter, ax=ax1)
cbar.set_label('Absolute Error (%)', fontsize=12, fontweight='bold')
ax1.legend(fontsize=12)
ax1.grid(True, alpha=0.3, linestyle='--')
plt.tight_layout()
plt.savefig('prediction_scatter_detailed.png', dpi=300, bbox_inches='tight')
print("Detailed scatter plot saved as 'prediction_scatter_detailed.png'")
plt.close()

# Graph 2: Large Bar Comparison (top 30 samples)
fig2, ax2 = plt.subplots(figsize=(16, 8))
n_bars = min(30, len(predictions_df))
indices = range(n_bars)
width = 0.35
bars1 = ax2.bar([i - width/2 for i in indices], 
                predictions_df['actual_removal_percentage (%)'].iloc[:n_bars], 
                width, label='Actual', alpha=0.8, color='royalblue', edgecolor='black')
bars2 = ax2.bar([i + width/2 for i in indices], 
                predictions_df['predicted_removal_percentage (%)'].iloc[:n_bars], 
                width, label='Predicted', alpha=0.8, color='darkorange', edgecolor='black')
ax2.set_xlabel('Sample Index', fontsize=14, fontweight='bold')
ax2.set_ylabel('Removal Percentage (%)', fontsize=14, fontweight='bold')
ax2.set_title(f'Detailed: Actual vs Predicted Values (First {n_bars} Samples)', 
              fontsize=16, fontweight='bold', pad=20)
ax2.legend(fontsize=12, loc='best')
ax2.set_xticks(indices)
ax2.set_xticklabels([str(i) for i in indices], rotation=45)
ax2.grid(True, alpha=0.3, axis='y', linestyle='--')
plt.tight_layout()
plt.savefig('prediction_bars_detailed.png', dpi=300, bbox_inches='tight')
print("Detailed bar chart saved as 'prediction_bars_detailed.png'")
plt.close()

print("\n✓ Created 3 graph files for All Predictions:")
print("  1. all_predictions_graphs.png (12 subplots overview)")
print("  2. prediction_scatter_detailed.png (detailed scatter plot)")
print("  3. prediction_bars_detailed.png (detailed bar comparison)")

print("\n" + "=" * 60)
print("All Visualizations Complete!")
print("=" * 60)
